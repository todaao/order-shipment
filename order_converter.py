"""
order_converter.py
発注試算Excelファイル → orders.json 変換スクリプト

使い方:
  python order_converter.py 発注試算_*.xlsx orders.json
  python order_converter.py 発注試算_A.xlsx 発注試算_B.xlsx orders.json

テンプレート構造（発注試算テンプレートシート）:
  B3=商品名, B4=ブランド, B5=発注予定日, B6=リードタイム週, B7=入庫予定日
  B8=容器MOQ, B9=バルクMOQ
  B13=toB月数, B14=EC月数, B15=海外月数

  ▼ 商品情報 SKU名/JAN（4+4二列レイアウト）:
    左ブロック(SKU1-4): E3:E6=SKU名, G3:G6=JAN
    右ブロック(SKU5-8): J3:J6=SKU名, L3:L6=JAN

  D18:K18 = 月ヘッダー（"N月" → 発注年と組み合わせてYYYY-MM）

  toB  SKU行: A20:A27, B=現在庫, D:K=月別数量
  EC   SKU行: A29:A36, B=現在庫, D:K=月別数量
  海外 SKU行: A38:A45, B=現在庫, D:K=月別数量

  確定発注数サマリー (rows 93-100):
    A=SKU名, B=JAN, C=現在庫, D=推計必要数, E=発注数, H=確定発注数

  Section C 小売ブロック (bases: 113,126,139,152,165,178,191,204,217,230):
    A{base}   = "[RETAILER] 名前"
    B{base+1} = 業態POS傾斜, D{base+1} = 店舗数, F{base+1} = 導入状況
    D:K{base+2} = 月別傾斜
    A{base+4}:A{base+11} = SKU名, B = P/S値, D:K = 月別数量
"""

import sys
import json
import glob
import re
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl が必要です: pip install openpyxl")

SHEET_NAME = "発注試算テンプレート"
COLS = list("DEFGHIJK")   # 8ヶ月分（D〜K列）
RETAILER_BASES = [113, 126, 139, 152, 165, 178, 191, 204, 217, 230]
SKIP_NAMES = {"（SKU5）", "（SKU6）", "（SKU7）", "（SKU8）", ""}


def v(ws, col, row):
    return ws[f"{col}{row}"].value

def s(ws, col, row):
    val = v(ws, col, row)
    return str(val).strip() if val is not None else ""

def n(ws, col, row):
    val = v(ws, col, row)
    try: return int(val) if val is not None else 0
    except: return 0

def f(ws, col, row):
    val = v(ws, col, row)
    try: return float(val) if val is not None else 0.0
    except: return 0.0


def date_str(val):
    if val is None: return None
    if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
    m = re.match(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", str(val).strip())
    return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else str(val).strip()


def month_keys(ws, order_date):
    """D18:K18 の "N月" ヘッダー → YYYY-MM リスト（8件）"""
    if not order_date:
        return [None] * 8
    base_year  = int(order_date[:4])
    base_month = int(order_date[5:7])
    keys = []
    for col in COLS:
        val = v(ws, col, 18)
        m = re.search(r"(\d{1,2})月", str(val)) if val else None
        if not m:
            keys.append(None)
            continue
        mo   = int(m.group(1))
        year = base_year if mo >= base_month else base_year + 1
        keys.append(f"{year}-{mo:02d}")
    return keys


def read_sku_info(ws):
    """
    商品情報部分の SKU名/JAN を 4+4 二列レイアウトから読み取る。
    左ブロック(SKU1-4): E列=SKU名, G列=JAN, rows 3-6
    右ブロック(SKU5-8): J列=SKU名, L列=JAN, rows 3-6
    戻り値: [(sku_name, jan), ...] インデックス順（最大8件）
    """
    result = []
    for r in range(3, 7):
        name = s(ws, "E", r)
        jan  = s(ws, "G", r)
        result.append((name, jan if jan else None))
    for r in range(3, 7):
        name = s(ws, "J", r)
        jan  = s(ws, "L", r)
        result.append((name, jan if jan else None))
    return result


def sku_rows(ws, r_start, r_end, mkeys):
    """A=SKU名, B=現在庫, D:K=月別数量 の行を読む"""
    result = {}
    for r in range(r_start, r_end + 1):
        name = s(ws, "A", r)
        if not name or name in SKIP_NAMES:
            continue
        monthly = {mk: n(ws, col, r) for col, mk in zip(COLS, mkeys) if mk}
        result[name] = {"stock": n(ws, "B", r), "monthly": monthly}
    return result


def parse_file(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        print(f"  [SKIP] シート '{SHEET_NAME}' なし: {path}")
        return None
    ws = wb[SHEET_NAME]

    product  = s(ws, "B", 3)
    brand    = s(ws, "B", 4)
    if not product:
        print(f"  [SKIP] 商品名が空: {path}")
        return None

    order_date   = date_str(v(ws, "B", 5))
    arrival_date = date_str(v(ws, "B", 7))
    mkeys        = month_keys(ws, order_date)

    # ── 商品情報部分からSKU名/JANを読む ──
    sku_info = read_sku_info(ws)  # [(name, jan), ...] × 8

    # ── チャネル別月別数量（A列のSKU名で紐づけ）──
    tob = sku_rows(ws, 20, 27, mkeys)
    ec  = sku_rows(ws, 29, 36, mkeys)
    ovs = sku_rows(ws, 38, 45, mkeys)

    # ── 確定発注数サマリー (A93:H100) ──
    # A=SKU名, B=JAN, C=現在庫, D=推計必要数, E=発注数, H=確定発注数
    summary = {}
    for r in range(93, 101):
        name = s(ws, "A", r)
        if not name or name in SKIP_NAMES:
            continue
        summary[name] = {
            "jan":       s(ws, "B", r) or None,
            "stock":     n(ws, "C", r),
            "estimated": n(ws, "D", r),
            "order_qty": n(ws, "E", r),
            "confirmed": n(ws, "H", r),
        }

    # ── SKUリストを構築 ──
    # 商品情報部分のSKU名を基準にし、確定発注数サマリーで補完
    all_names = [name for name, jan in sku_info if name]
    # サマリーにあってsKU情報にない名前も追加
    for name in summary:
        if name not in all_names:
            all_names.append(name)
    # toB/EC/海外にしかない名前も追加
    for d in [tob, ec, ovs]:
        for name in d:
            if name not in all_names:
                all_names.append(name)

    # JAN: 商品情報部分 → 確定発注数サマリー の優先順
    jan_by_name = {name: jan for name, jan in sku_info if name and jan}
    for name, info in summary.items():
        if info.get("jan") and name not in jan_by_name:
            jan_by_name[name] = info["jan"]

    skus = []
    for name in all_names:
        info  = summary.get(name, {})
        stock = info.get("stock") or tob.get(name, {}).get("stock", 0)
        skus.append({
            "sku_name":        name,
            "jan":             jan_by_name.get(name),
            "current_stock":   stock,
            "estimated_total": info.get("estimated", 0),
            "order_quantity":  info.get("order_qty", 0),
            "confirmed_order": info.get("confirmed", 0),
            "monthly_summary": {
                "tob":      tob.get(name, {}).get("monthly", {}),
                "ec":       ec.get(name,  {}).get("monthly", {}),
                "overseas": ovs.get(name, {}).get("monthly", {}),
            },
        })

    # ── Section C 小売ブロック ──
    retailer_blocks = []
    for base in RETAILER_BASES:
        header = s(ws, "A", base)
        if not header.startswith("[RETAILER]"):
            continue
        rname = re.sub(r"\[RETAILER\]|★新規", "", header).strip()
        if not rname:
            continue
        mt = {mk: f(ws, col, base+2) for col, mk in zip(COLS, mkeys) if mk}
        sku_ps, sku_mo = {}, {}
        for r in range(base+4, base+12):
            sn = s(ws, "A", r)
            if not sn or sn in SKIP_NAMES:
                continue
            sku_ps[sn] = f(ws, "B", r)
            sku_mo[sn] = {mk: n(ws, col, r) for col, mk in zip(COLS, mkeys) if mk}
        retailer_blocks.append({
            "retailer":   rname,
            "status":     s(ws, "F", base+1) or "導入済み",
            "posTilt":    f(ws, "B", base+1),
            "stores":     n(ws, "D", base+1),
            "monthTilts": mt,
            "skuPs":      sku_ps,
            "skus":       sku_mo,
        })

    return {
        "order_id":        Path(path).stem,
        "brand":           brand,
        "product":         product,
        "order_date":      order_date,
        "arrival_date":    arrival_date,
        "lead_time_weeks": n(ws, "B", 6),
        "container_moq":   n(ws, "B", 8),
        "bulk_moq":        n(ws, "B", 9),
        "tob_months":      n(ws, "B", 13),
        "ec_months":       n(ws, "B", 14),
        "overseas_months": n(ws, "B", 15),
        "skus":            skus,
        "retailer_blocks": retailer_blocks,
    }


def main():
    args = sys.argv[1:]
    if len(args) < 2:
        sys.exit("使い方: python order_converter.py 発注試算_*.xlsx orders.json")

    output_path = args[-1]
    input_files = []
    for pat in args[:-1]:
        expanded = sorted(glob.glob(pat))
        input_files.extend(expanded if expanded else ([pat] if Path(pat).exists() else []))

    if not input_files:
        sys.exit("[ERROR] 入力ファイルが見つかりません")

    print(f"対象: {len(input_files)} ファイル")
    orders = []
    for fp in input_files:
        print(f"  読み込み: {fp}")
        result = parse_file(fp)
        if result:
            orders.append(result)
            sku_with_jan = sum(1 for s in result['skus'] if s.get('jan'))
            print(f"    → {result['brand']} / {result['product']} "
                  f"({len(result['skus'])} SKU / JAN取得 {sku_with_jan}件, "
                  f"{len(result['retailer_blocks'])} 小売)")

    with open(output_path, "w", encoding="utf-8") as fout:
        json.dump({"orders": orders}, fout, ensure_ascii=False, indent=2)
    print(f"\n出力: {output_path}  ({len(orders)} 発注)")


if __name__ == "__main__":
    main()
